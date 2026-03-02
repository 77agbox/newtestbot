import asyncio
import logging
import os
import json
from aiogram import Bot, Dispatcher, F
from aiogram.enums import ParseMode
from aiogram.types import (
    Message,
    CallbackQuery,
    InlineKeyboardMarkup,
    InlineKeyboardButton
)
from aiogram.client.default import DefaultBotProperties
from aiogram.filters import CommandStart, Command
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from openpyxl import load_workbook

# ================= CONFIG =================

BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID"))

logging.basicConfig(level=logging.INFO)

bot = Bot(
    token=BOT_TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML)
)

dp = Dispatcher(storage=MemoryStorage())

MASTER_FILE = "masterclasses.json"

# ================= FSM =================

class ClubForm(StatesGroup):
    age = State()
    address = State()
    direction = State()
    clubs = State()

class PackageForm(StatesGroup):
    people = State()
    activities = State()
    name = State()
    phone = State()

class MasterForm(StatesGroup):
    title = State()
    description = State()
    date = State()
    price = State()
    teacher = State()
    link = State()

    enroll_name = State()
    enroll_phone = State()
    enroll_index = State()

class SupportForm(StatesGroup):
    text = State()

# ================= UTIL =================

def profile_link(user):
    return (
        f'<a href="https://t.me/{user.username}">@{user.username}</a>'
        if user.username else f'<a href="tg://user?id={user.id}">Профиль</a>'
    )

def parse_age_range(age_text: str):
    if not age_text:
        return None, None
    text = age_text.lower().replace("лет", "").replace(" ", "")
    if "-" in text:
        a, b = text.split("-")
        if a.isdigit() and b.isdigit():
            return int(a), int(b)
    if "+" in text:
        num = text.replace("+", "")
        if num.isdigit():
            return int(num), 99
    if text.isdigit():
        age = int(text)
        return age, age
    return None, None

def load_clubs():
    wb = load_workbook("joined_clubs.xlsx")
    sheet = wb.active
    clubs = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        clubs.append({
            "direction": row[0],
            "name": row[1],
            "age": row[2],
            "address": row[3],
            "teacher": row[4],
            "link": row[5],
        })
    return clubs

def load_masterclasses():
    if not os.path.exists(MASTER_FILE):
        with open(MASTER_FILE, "w", encoding="utf-8") as f:
            json.dump([], f)
        return []
    with open(MASTER_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_masterclasses(data):
    with open(MASTER_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

# ================= KEYBOARDS =================

def main_menu(user_id):
    buttons = [
        [InlineKeyboardButton(text="🎨 Кружки", callback_data="clubs")],
        [InlineKeyboardButton(text="🧩 Мастер-классы", callback_data="masters")],
        [InlineKeyboardButton(text="🎉 Пакетные туры", callback_data="packages")],
        [InlineKeyboardButton(text="✉ Написать в поддержку", callback_data="support")]
    ]
    if user_id == ADMIN_ID:
        buttons.append([InlineKeyboardButton(text="⚙ Админ панель", callback_data="admin")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

# ================= START =================

@dp.message(CommandStart())
async def start(message: Message):
    add_user(message.from_user.id)

    await message.answer(
        "Приветствую! Я Бот Виктор!\n"
        "Я помогу вам выбрать интересные занятия в нашем центре.\n\n"
        "Выберите раздел:",
        reply_markup=main_menu(message.from_user.id)
    )

# ================= MENU =================

@dp.callback_query(F.data == "menu")
async def menu(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_text(
        "Главное меню:",
        reply_markup=main_menu(callback.from_user.id)
    )
    await callback.answer()

# ================= SUPPORT =================

@dp.callback_query(F.data == "support")
async def support_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(SupportForm.text)
    await callback.message.answer("Напишите ваше сообщение:")
    await callback.answer()

@dp.message(SupportForm.text)
async def support_send(message: Message, state: FSMContext):
    await bot.send_message(
        ADMIN_ID,
        f"✉ Поддержка\n\n"
        f"Профиль: {profile_link(message.from_user)}\n"
        f"{message.text}",
        disable_web_page_preview=True
    )
    await message.answer("Сообщение отправлено администратору ✅")
    await state.clear()

# ================= CLUBS 3.0 =================

@dp.callback_query(F.data == "clubs")
async def clubs_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(ClubForm.age)
    await callback.message.edit_text("Укажите возраст:")
    await callback.answer()


@dp.message(ClubForm.age)
async def clubs_age(message: Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("Введите возраст числом.")
        return

    await state.update_data(age=int(message.text))
    await state.set_state(ClubForm.address)

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Главное здание", callback_data="addr_0")],
        [InlineKeyboardButton(text="МХС Аннино", callback_data="addr_1")],
        [InlineKeyboardButton(text="СП Юный техник", callback_data="addr_2")],
        [InlineKeyboardButton(text="СП Щербинка", callback_data="addr_3")],
        [InlineKeyboardButton(text="Онлайн", callback_data="addr_4")],
        [InlineKeyboardButton(text="⬅ В меню", callback_data="menu")]
    ])

    await message.answer("Выберите подразделение:", reply_markup=keyboard)


@dp.callback_query(F.data.startswith("addr_"))
async def clubs_address(callback: CallbackQuery, state: FSMContext):
    index = int(callback.data.split("_")[1])
    data = await state.get_data()
    clubs = load_clubs()

    address_filters = [
        "газопровод",
        "варшав",
        "нагатин",
        "пушкин",
        ""  # онлайн
    ]

    filtered = []

    for club in clubs:
        min_age, max_age = parse_age_range(str(club["age"]))
        if min_age is None:
            continue

        if not (min_age <= data["age"] <= max_age):
            continue

        address = str(club["address"]).lower()

        if index == 4:
            if not address.strip():
                filtered.append(club)
        else:
            if address_filters[index] in address:
                filtered.append(club)

    if not filtered:
        await callback.message.answer("Подходящих кружков не найдено.")
        await state.clear()
        await callback.answer()
        return

    directions = sorted(set(c["direction"] for c in filtered))
    await state.update_data(clubs=filtered)

    buttons = [
        [InlineKeyboardButton(text=d, callback_data=f"dir_{i}")]
        for i, d in enumerate(directions)
    ]
    buttons.append([InlineKeyboardButton(text="⬅ В меню", callback_data="menu")])

    await callback.message.answer(
        "Выберите направление:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )

    await state.set_state(ClubForm.direction)
    await callback.answer()


@dp.callback_query(F.data.startswith("dir_"))
async def clubs_direction(callback: CallbackQuery, state: FSMContext):
    index = int(callback.data.split("_")[1])
    data = await state.get_data()

    clubs = data["clubs"]
    directions = sorted(set(c["direction"] for c in clubs))

    if index >= len(directions):
        await callback.answer("Ошибка выбора", show_alert=True)
        return

    selected_direction = directions[index]
    result = [c for c in clubs if c["direction"] == selected_direction]

    await state.update_data(clubs=result)

    buttons = [
        [InlineKeyboardButton(text=c["name"], callback_data=f"club_{i}")]
        for i, c in enumerate(result)
    ]
    buttons.append([InlineKeyboardButton(text="⬅ Назад", callback_data="clubs")])
    buttons.append([InlineKeyboardButton(text="⬅ В меню", callback_data="menu")])

    await callback.message.answer(
        "Выберите кружок:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )

    await callback.answer()


@dp.callback_query(F.data.startswith("club_"))
async def club_card(callback: CallbackQuery, state: FSMContext):
    index = int(callback.data.split("_")[1])
    data = await state.get_data()

    clubs = data["clubs"]

    if index >= len(clubs):
        await callback.answer("Ошибка выбора", show_alert=True)
        return

    club = clubs[index]

    text = (
        f"<b>{club['name']}</b>\n\n"
        f"Возраст: {club['age']}\n"
        f"Педагог: {club['teacher']}\n"
        f"Адрес: {club['address']}\n\n"
        f"<a href='{club['link']}'>Перейти к записи</a>"
    )

    await callback.message.answer(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅ В меню", callback_data="menu")]
        ])
    )

    await callback.answer()

# ================= MASTERCLASSES 3.0 =================

# ---------- Пользователь ----------

@dp.callback_query(F.data == "masters")
async def masters_list(callback: CallbackQuery):
    masters = load_masterclasses()

    if not masters:
        await callback.message.answer("Мастер-классы пока не добавлены.")
        await callback.answer()
        return

    buttons = [
        [InlineKeyboardButton(text=m["title"], callback_data=f"master_{i}")]
        for i, m in enumerate(masters)
    ]
    buttons.append([InlineKeyboardButton(text="⬅ В меню", callback_data="menu")])

    await callback.message.answer(
        "Доступные мастер-классы:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )
    await callback.answer()


@dp.callback_query(F.data.startswith("master_"))
async def master_card(callback: CallbackQuery):
    index = int(callback.data.split("_")[1])
    masters = load_masterclasses()

    if index >= len(masters):
        await callback.answer("Ошибка выбора", show_alert=True)
        return

    m = masters[index]

    text = (
        f"━━━━━━━━━━━━━━━\n"
        f"🎨 <b>{m['title']}</b>\n"
        f"━━━━━━━━━━━━━━━\n\n"
        f"📝 <b>Описание:</b>\n"
        f"{m['description']}\n\n"
        f"📅 <b>Дата и время:</b> {m['date']}\n"
        f"💰 <b>Стоимость:</b> {m['price']} ₽\n"
        f"👩‍🏫 <b>Педагог:</b> {m['teacher']}\n\n"
        f"🔗 <a href='{m['link']}'>Подробнее</a>\n\n"
        f"━━━━━━━━━━━━━━━"
    )

    await callback.message.answer(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✉ Записаться", callback_data=f"enroll_{index}")],
            [InlineKeyboardButton(text="⬅ Назад", callback_data="masters")],
            [InlineKeyboardButton(text="⬅ В меню", callback_data="menu")]
        ])
    )

    await callback.answer()


# ---------- Запись на МК с вводом данных ----------

@dp.callback_query(F.data.startswith("enroll_"))
async def master_enroll_start(callback: CallbackQuery, state: FSMContext):
    index = int(callback.data.split("_")[1])

    masters = load_masterclasses()
    if index >= len(masters):
        await callback.answer("Ошибка", show_alert=True)
        return

    await state.update_data(enroll_index=index)
    await state.set_state(MasterForm.enroll_name)

    await callback.message.answer("Как к вам обращаться?")
    await callback.answer()


@dp.message(MasterForm.enroll_name)
async def master_enroll_name(message: Message, state: FSMContext):
    await state.update_data(enroll_name=message.text.strip())
    await state.set_state(MasterForm.enroll_phone)
    await message.answer("Введите номер телефона для связи:")


@dp.message(MasterForm.enroll_phone)
async def master_enroll_finish(message: Message, state: FSMContext):
    data = await state.get_data()
    masters = load_masterclasses()

    index = data["enroll_index"]
    if index >= len(masters):
        await message.answer("Ошибка.")
        await state.clear()
        return

    m = masters[index]
    name = data["enroll_name"]
    phone = message.text.strip()

    await bot.send_message(
        ADMIN_ID,
        f"📚 <b>Новая запись на мастер-класс</b>\n\n"
        f"<b>{m['title']}</b>\n\n"
        f"👤 Имя: {name}\n"
        f"📞 Телефон: {phone}\n\n"
        f"Профиль: {profile_link(message.from_user)}\n"
        f"TG ID: {message.from_user.id}",
        disable_web_page_preview=True
    )

    await message.answer("Заявка отправлена администратору ✅")
    await state.clear()

# ---------- Админ ----------

@dp.callback_query(F.data == "admin")
async def admin_panel(callback: CallbackQuery):
    if callback.from_user.id != ADMIN_ID:
        await callback.answer("Нет доступа", show_alert=True)
        return

    await callback.message.answer(
        "Админ панель — Мастер-классы:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="➕ Добавить МК", callback_data="add_master")],
            [InlineKeyboardButton(text="❌ Удалить МК", callback_data="delete_master")],
            [InlineKeyboardButton(text="⬅ В меню", callback_data="menu")]
        ])
    )
    await callback.answer()


@dp.callback_query(F.data == "add_master")
async def master_add_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(MasterForm.title)
    await callback.message.answer("Введите название мастер-класса:")
    await callback.answer()


@dp.message(MasterForm.title)
async def master_title(message: Message, state: FSMContext):
    await state.update_data(title=message.text)
    await state.set_state(MasterForm.description)
    await message.answer("Введите описание:")


@dp.message(MasterForm.description)
async def master_description(message: Message, state: FSMContext):
    await state.update_data(description=message.text)
    await state.set_state(MasterForm.date)
    await message.answer("Введите дату и время:")


@dp.message(MasterForm.date)
async def master_date(message: Message, state: FSMContext):
    await state.update_data(date=message.text)
    await state.set_state(MasterForm.price)
    await message.answer("Введите стоимость:")


@dp.message(MasterForm.price)
async def master_price(message: Message, state: FSMContext):
    await state.update_data(price=message.text)
    await state.set_state(MasterForm.teacher)
    await message.answer("Введите педагога:")


@dp.message(MasterForm.teacher)
async def master_teacher(message: Message, state: FSMContext):
    await state.update_data(teacher=message.text)
    await state.set_state(MasterForm.link)
    await message.answer("Введите ссылку на подробное описание:")


@dp.message(MasterForm.link)
async def master_save(message: Message, state: FSMContext):
    data = await state.get_data()
    data["link"] = message.text

    masters = load_masterclasses()
    masters.append(data)
    save_masterclasses(masters)

    await message.answer("Мастер-класс добавлен ✅")
    await state.clear()


@dp.callback_query(F.data == "delete_master")
async def master_delete_list(callback: CallbackQuery):
    masters = load_masterclasses()

    if not masters:
        await callback.answer("Нет мастер-классов", show_alert=True)
        return

    buttons = [
        [InlineKeyboardButton(text=f"❌ {m['title']}", callback_data=f"del_{i}")]
        for i, m in enumerate(masters)
    ]

    await callback.message.answer(
        "Выберите МК для удаления:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )
    await callback.answer()


@dp.callback_query(F.data.startswith("del_"))
async def master_delete_confirm(callback: CallbackQuery):
    index = int(callback.data.split("_")[1])
    masters = load_masterclasses()

    if index < len(masters):
        masters.pop(index)
        save_masterclasses(masters)

    await callback.answer("Удалено ✅", show_alert=True)

# ================= PACKAGES 3.0 =================

PACKAGE_MODULES = {
    "Картинг": [2200, 2100, 2000],
    "Симрейсинг": [1600, 1500, 1400],
    "Практическая стрельба": [1600, 1500, 1400],
    "Лазертаг": [1600, 1500, 1400],
    "Керамика": [1600, 1500, 1400],
    "Мягкая игрушка": [1300, 1200, 1100],
}


def activities_keyboard(selected=None):
    selected = selected or []
    buttons = []

    for i, name in enumerate(PACKAGE_MODULES.keys()):
        prefix = "✅ " if name in selected else ""
        buttons.append([
            InlineKeyboardButton(
                text=f"{prefix}{name}",
                callback_data=f"act_{i}"
            )
        ])

    buttons.append([InlineKeyboardButton(text="🟢 Готово", callback_data="act_done")])
    buttons.append([InlineKeyboardButton(text="⬅ В меню", callback_data="menu")])

    return InlineKeyboardMarkup(inline_keyboard=buttons)


@dp.callback_query(F.data == "packages")
async def package_start(callback: CallbackQuery, state: FSMContext):
    await state.set_state(PackageForm.people)
    await callback.message.answer(
        "👥 Введите количество человек (минимум 5):"
    )
    await callback.answer()


@dp.message(PackageForm.people)
async def package_people(message: Message, state: FSMContext):
    if not message.text.isdigit() or int(message.text) < 5:
        await message.answer("Минимум 5 человек.")
        return

    await state.update_data(
        people=int(message.text),
        selected=[]
    )
    await state.set_state(PackageForm.activities)

    await message.answer(
        "🎯 Выберите от 1 до 3 активностей:",
        reply_markup=activities_keyboard()
    )


@dp.callback_query(F.data.startswith("act_"))
async def package_choose_activity(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected = data.get("selected", [])

    if callback.data == "act_done":
        if not 1 <= len(selected) <= 3:
            await callback.answer("Выберите 1–3 активности", show_alert=True)
            return

        await state.set_state(PackageForm.name)
        await callback.message.answer("Введите ваше имя:")
        await callback.answer()
        return

    index = int(callback.data.split("_")[1])
    activity = list(PACKAGE_MODULES.keys())[index]

    if activity in selected:
        selected.remove(activity)
    else:
        if len(selected) >= 3:
            await callback.answer("Можно выбрать максимум 3 активности", show_alert=True)
            return
        selected.append(activity)

    await state.update_data(selected=selected)

    # обновляем клавиатуру с галочками
    await callback.message.edit_reply_markup(
        reply_markup=activities_keyboard(selected)
    )
    await callback.answer()


@dp.message(PackageForm.name)
async def package_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text.strip())
    await state.set_state(PackageForm.phone)
    await message.answer("📞 Введите телефон для связи:")


@dp.message(PackageForm.phone)
async def package_finish(message: Message, state: FSMContext):
    data = await state.get_data()

    people = data["people"]
    selected = data["selected"]
    name = data["name"]
    phone = message.text.strip()

    price_index = len(selected) - 1

    total = 0
    per_person_total = 0
    lines = []

    for act in selected:
        price = PACKAGE_MODULES[act][price_index]
        cost = price * people
        total += cost
        per_person_total += price
        lines.append(f"• {act}: <b>{price} ₽</b> с человека")

    activities_text = "\n".join(lines)

    await bot.send_message(
        ADMIN_ID,
        f"🎉 Новая заявка на пакетный тур\n\n"
        f"Клиент: {name}\n"
        f"Телефон: {phone}\n"
        f"Профиль: {profile_link(message.from_user)}\n"
        f"TG ID: {message.from_user.id}\n\n"
        f"Группа: {people} человек\n"
        f"Активности: {', '.join(selected)}\n\n"
        f"{activities_text}\n\n"
        f"С человека: {per_person_total} ₽\n"
        f"Общая сумма: {total} ₽",
        disable_web_page_preview=True
    )

    await message.answer(
        f"✅ Ваша заявка принята!\n\n"
        f"{activities_text}\n\n"
        f"💰 С человека: {per_person_total} ₽\n"
        f"👥 Общая сумма: {total} ₽",
    )

    await state.clear()

# ================= PRO BROADCAST SYSTEM =================

USERS_FILE = "users.json"


def load_users():
    if not os.path.exists(USERS_FILE):
        return []
    with open(USERS_FILE, "r") as f:
        return json.load(f)


def save_users(users):
    with open(USERS_FILE, "w") as f:
        json.dump(users, f)


def add_user(user_id):
    users = load_users()
    if user_id not in users:
        users.append(user_id)
        save_users(users)


def remove_user(user_id):
    users = load_users()
    if user_id in users:
        users.remove(user_id)
        save_users(users)


# ====== ДОБАВЬТЕ В /start add_user ======
# Внутри вашего start обязательно должна быть строка:
# add_user(message.from_user.id)


class BroadcastForm(StatesGroup):
    content = State()


# -------- Команда статистики --------

@dp.message(Command("users"))
async def users_stat(message: Message):
    if message.from_user.id != ADMIN_ID:
        return

    users = load_users()
    await message.answer(f"👥 Всего пользователей: {len(users)}")


# -------- Запуск рассылки --------

@dp.message(Command("broadcast"))
async def broadcast_start(message: Message, state: FSMContext):
    if message.from_user.id != ADMIN_ID:
        return

    await state.set_state(BroadcastForm.content)
    await message.answer(
        "Отправьте текст или фото с подписью для рассылки."
    )


# -------- Отправка рассылки --------

@dp.message(BroadcastForm.content)
async def broadcast_send(message: Message, state: FSMContext):
    users = load_users()
    success = 0
    removed = 0

    unsubscribe_kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(
                text="ℹ Управление уведомлениями",
                callback_data="manage_notifications"
            )]
        ]
    )

    for user in users.copy():
        user_id = user["id"] if isinstance(user, dict) else user

        try:
            if message.photo:
                photo = message.photo[-1].file_id
                await bot.send_photo(
                    user_id,
                    photo,
                    caption=message.caption or "",
                    reply_markup=unsubscribe_kb
                )
            else:
                await bot.send_message(
                    user_id,
                    message.text,
                    reply_markup=unsubscribe_kb
                )

            success += 1
            await asyncio.sleep(0.05)

        except Exception as e:
            if "bot was blocked" in str(e):
                remove_user(user_id)
                removed += 1
            continue

    await message.answer(
        f"📊 Рассылка завершена\n\n"
        f"✅ Доставлено: {success}\n"
        f"🚫 Удалено (заблокировали): {removed}"
    )

    await state.clear()


# -------- Управление уведомлениями --------

@dp.callback_query(F.data == "manage_notifications")
async def manage_notifications(callback: CallbackQuery):
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(
                text="🔕 Отписаться от рассылки",
                callback_data="unsubscribe_confirm"
            )],
            [InlineKeyboardButton(
                text="⬅ Назад",
                callback_data="close_manage"
            )]
        ]
    )

    await callback.message.edit_reply_markup(reply_markup=kb)
    await callback.answer()


@dp.callback_query(F.data == "unsubscribe_confirm")
async def unsubscribe_confirm(callback: CallbackQuery):
    remove_user(callback.from_user.id)

    await callback.message.edit_text(
        "🔕 Вы отписались от рассылки.\n\n"
        "Чтобы снова получать уведомления — нажмите /start"
    )
    await callback.answer()


@dp.callback_query(F.data == "close_manage")
async def close_manage(callback: CallbackQuery):
    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.answer()

# ================= RUN =================

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
