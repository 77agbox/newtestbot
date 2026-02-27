from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram import Router
from openpyxl import load_workbook

router = Router()

# Загрузка данных кружков из Excel
def load_clubs():
    wb = load_workbook('joined_clubs.xlsx')
    sheet = wb.active
    clubs = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        clubs.append(row)
    return clubs

@router.message(Command("clubs"))
async def show_clubs(message: types.Message):
    clubs = load_clubs()
    text = "Доступные кружки:\n"
    for club in clubs:
        text += f"{club[1]} (возраст: {club[2]}) - {club[3]}\n"
    await message.answer(text)
